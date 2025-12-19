import cv2
import os
from datetime import datetime
import calendar
from openpyxl import Workbook, load_workbook
import streamlit as st
import numpy as np
from PIL import Image
import time

# ---------------- PAGE SETUP ----------------
st.set_page_config(page_title="Face Attendance System", layout="centered")
st.title("😊 Good Morning! Have a Nice Day")
st.write("Face Detection Attendance System")

# ---------------- SESSION STATE ----------------
if "marked" not in st.session_state:
    st.session_state["marked"] = False
if "camera_running" not in st.session_state:
    st.session_state["camera_running"] = False
if "face_detected_count" not in st.session_state:
    st.session_state["face_detected_count"] = 0

# ---------------- EXCEL FILE ----------------
excel_file = "Attendance.xlsx"

if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name", "Date", "Day", "Time"])
    wb.save(excel_file)

# ---------------- FACE DETECTOR (HAAR CASCADE) ----------------
@st.cache_resource
def load_face_detector():
    """Load and cache the face detector"""
    cascade_path = cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
    return cv2.CascadeClassifier(cascade_path)

face_cascade = load_face_detector()

# ---------------- ATTENDANCE FUNCTION ----------------
def mark_attendance(name):
    """Mark attendance for a person"""
    try:
        now = datetime.now()
        date = now.strftime("%Y-%m-%d")
        day = calendar.day_name[now.weekday()]
        time = now.strftime("%H:%M:%S")

        wb = load_workbook(excel_file)
        ws = wb.active

        # Check if attendance already marked today
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == name and row[1] == date:
                return False, "Already marked today"

        ws.append([name, date, day, time])
        wb.save(excel_file)
        return True, "Attendance marked successfully"
    except Exception as e:
        return False, f"Error: {str(e)}"

# ---------------- FACE DETECTION FUNCTION ----------------
def detect_faces(frame, name):
    """Detect faces in the frame and draw rectangles"""
    # Convert to grayscale for detection
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    
    # Detect faces
    faces = face_cascade.detectMultiScale(
        gray,
        scaleFactor=1.1,
        minNeighbors=5,
        minSize=(30, 30),
        flags=cv2.CASCADE_SCALE_IMAGE
    )
    
    # Draw rectangles around faces
    for (x, y, w, h) in faces:
        # Draw rectangle
        cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
        
        # Add name label
        label = name if name else "Unknown"
        cv2.putText(frame, label, (x, y-10), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
        
        # Add "Face Detected" text
        cv2.putText(frame, "Face Detected!", (x, y+h+25), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
    
    return frame, len(faces)

# ---------------- STREAMLIT UI ----------------
# Sidebar for viewing attendance
with st.sidebar:
    st.header("📊 Attendance Records")
    if st.button("View Today's Attendance"):
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            today = datetime.now().strftime("%Y-%m-%d")
            
            st.subheader(f"Attendance for {today}")
            today_records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] == today:
                    today_records.append(row)
            
            if today_records:
                import pandas as pd
                df = pd.DataFrame(today_records, columns=["Name", "Date", "Day", "Time"])
                st.dataframe(df, width="stretch")
            else:
                st.info("No attendance records for today")
        except Exception as e:
            st.error(f"Error reading attendance: {str(e)}")
    
    if st.button("View All Records"):
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            
            all_records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Check if name exists
                    all_records.append(row)
            
            if all_records:
                import pandas as pd
                df = pd.DataFrame(all_records, columns=["Name", "Date", "Day", "Time"])
                st.dataframe(df, width="stretch")
                st.info(f"Total records: {len(all_records)}")
            else:
                st.info("No attendance records yet")
        except Exception as e:
            st.error(f"Error reading attendance: {str(e)}")

# Main content
st.divider()

# Name input
name = st.text_input("📝 Enter Your Name", placeholder="John Doe")

# Camera controls
col1, col2 = st.columns([1, 1])
with col1:
    start_camera = st.button("🎥 Start Camera", type="primary")
with col2:
    stop_camera = st.button("⏹️ Stop Camera")

if start_camera and name.strip():
    st.session_state["camera_running"] = True
    st.session_state["marked"] = False
    st.session_state["face_detected_count"] = 0
elif start_camera and not name.strip():
    st.warning("⚠️ Please enter your name first!")
    st.session_state["camera_running"] = False

if stop_camera:
    st.session_state["camera_running"] = False
    st.session_state["marked"] = False

# Camera feed and face detection
if st.session_state.get("camera_running", False) and name.strip():
    st.info("📷 Opening camera... Position your face clearly in the frame.")
    
    # Placeholder for camera feed
    camera_placeholder = st.empty()
    status_placeholder = st.empty()
    
    # Try to open camera
    cap = cv2.VideoCapture(0)
    
    if not cap.isOpened():
        st.error("❌ Could not access camera. Please check camera permissions.")
        st.session_state["camera_running"] = False
    else:
        st.success("✅ Camera started successfully!")
        
        # Frame counter for marking attendance
        consecutive_detections = 0
        required_detections = 10  # Need 10 consecutive detections to mark
        
        # Process frames while camera is running
        while st.session_state.get("camera_running", False):
            ret, frame = cap.read()
            
            if not ret:
                st.error("❌ Failed to grab frame from camera")
                break
            
            # Detect faces and draw rectangles
            processed_frame, num_faces = detect_faces(frame, name)
            
            # Convert BGR to RGB for display
            frame_rgb = cv2.cvtColor(processed_frame, cv2.COLOR_BGR2RGB)
            
            # Display the frame
            camera_placeholder.image(frame_rgb, channels="RGB", use_container_width=True)
            
            # Handle face detection and attendance marking
            if num_faces > 0:
                consecutive_detections += 1
                st.session_state["face_detected_count"] = consecutive_detections
                
                status_placeholder.info(f"👤 Face detected! ({consecutive_detections}/{required_detections})")
                
                # Mark attendance after consistent detection
                if consecutive_detections >= required_detections and not st.session_state.get("marked", False):
                    success, message = mark_attendance(name)
                    st.session_state["marked"] = True
                    
                    if success:
                        status_placeholder.success(f"✅ Attendance marked for {name}!")
                        st.balloons()
                        time.sleep(2)
                        st.session_state["camera_running"] = False
                        break
                    else:
                        status_placeholder.warning(f"⚠️ {message}")
                        time.sleep(2)
                        st.session_state["camera_running"] = False
                        break
            else:
                consecutive_detections = 0
                st.session_state["face_detected_count"] = 0
                status_placeholder.warning("⏳ No face detected. Please position yourself clearly.")
            
            # Small delay to control frame rate
            time.sleep(0.1)
        
        # Release camera
        cap.release()
        
        if not st.session_state.get("camera_running", False):
            camera_placeholder.empty()
            if not st.session_state.get("marked", False):
                status_placeholder.info("Camera stopped.")

# Footer
st.divider()
st.caption("💡 Tips:")
st.caption("• Make sure your face is clearly visible and well-lit")
st.caption("• Look directly at the camera")
st.caption("• Remove any obstructions (sunglasses, masks)")
st.caption("• Stay still for a few seconds when face is detected")
